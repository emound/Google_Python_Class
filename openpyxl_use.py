import openpyxl

inv_file=openpyxl.load_workbook("inventory.xlsx")
product_list=inv_file["Sheet1"]

products_per_supplier={}
total_value_per_supplier={}
product_under_10_inv={}
product_list.cell(1,5).value="TotalPrice"


#range is exclusive of last 
for products_row in range(2, product_list.max_row+1):
    #parameter for cell() is row, column to access the cell; this accesses the cell not the value; .value will access the actual value
    # for dict. more recommended way of getting the value is dict.get(keyname) and to set user dict[keyname]
    supplier_name=product_list.cell(products_row, 4).value
    inventory=product_list.cell(products_row,2).value
    price=product_list.cell(products_row,3).value
    product_num=product_list.cell(products_row,1).value
    inventory_price=product_list.cell(products_row,5)


    #calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        current_number_of_prod=products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name]=current_number_of_prod+1
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name]=1

    #calculate total value per supplier
    if supplier_name in total_value_per_supplier:
        current_value=total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name]=(inventory*price) + current_value
    
    else:
        total_value_per_supplier[supplier_name]=inventory*price

    
    #inventory less than 10 for a specific product
    if inventory < 10:
        product_under_10_inv[product_num]=inventory

    #create new column and add the total price = inventory*price
    inventory_price.value=inventory*price



print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)

inv_file.save("UpdatedInventory.xlsx")
