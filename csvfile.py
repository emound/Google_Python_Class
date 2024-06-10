import csv
import openpyxl
with open('names.csv', 'r') as csvfile:
    csv_reader=csv.reader(csvfile)
    for line in csv_reader:
        print(line)


#load the workbook
excelfile=openpyxl.load_workbook("names.xlsx")
#Specifying the sheet
sheet1=excelfile["names"]

#printing all values from column1
for rows in range(1,sheet1.max_row+1):
    print(sheet1.cell(rows,1).value)
    

